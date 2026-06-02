#!/usr/bin/python3
"""
Map Helper - Parse linker map file and generate size statistics.

This script parses linker-generated .map files and generates detailed
size statistics reports in multiple formats (TXT, CSV, JSON, Excel).

Features:
- Parse ARM GCC linker map files (.map)
- Generate size statistics (Flash, RAM, Code, Rodata, Data, BSS, etc.)
- Support multiple output formats (TXT, CSV, JSON, Excel)
- Excel output has multiple sheets: Total, Detail, Top15 visualizations, plus paired Excel sheets sorted by Flash or by library name with rankings
- Top15 visualization: Top 15 libraries by Text/BSS/Data/RAM with charts
- Compare two map files to see size differences with diff highlighting
- Sorting and filtering capabilities
- Summary statistics with percentages
- Process single or multiple map files at once with consistent basename outputs
- Automatic merged Excel generation for multiple files
"""

import re
import os
import sys
import csv
import json
import argparse
from collections import OrderedDict
from typing import Dict, Optional, Tuple, List

# Try to import openpyxl for Excel support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class MapHelper(object):
    """
    Parse linker map file and generate size statistics.

    This class provides functionality to:
    - Parse ARM GCC linker-generated .map files
    - Extract size information for libraries and object files
    - Generate reports in multiple formats (TXT, CSV, JSON, Excel)
    - Generate Top15 hotspot visualization: Top 15 libraries by Text/BSS/Data/RAM with visualization charts
    - Compare two map files to identify size differences
    - Calculate summary statistics with percentages
    """

    # Column width for formatted output
    COLUMN_WIDTH = 7

    # CSV column headers
    CSV_COLUMNS = ['flash', 'ram', 'text', 'code', 'rodata', 'data', 'bss', 'dec', 'hex', 'filename', 'library_path']
    CSV_DETAIL_COLUMNS = ['flash', 'ram', 'text', 'code', 'rodata', 'data', 'bss', 'dec', 'hex', 'obj_name', 'filename', 'library_path']

    # Section type flags mapping
    SECTION_FLAGS = {
        'rodata': ['.rodata.', '.rodata', '.srodata.', '.vectors', '.vectors_itcm'],
        'text': ['.text.', '.text', '.stext.', '.itcm_sec_code', 'itcm_section', '.itcm', '.iram', '.interrupt'],
        'data': ['.data.', '.data', '.sdata.', '.dtcm_sec_data', '.dtcm'],
        'bss': ['.bss.', '.bss', '.sbss.', '.dtcm_sec_bss', '.bt_spec_data', '.ble_bss_data', 'video_spec_data'],
    }

    def __init__(self, map_file_path: str, min_size: int = 0, sort_by: str = 'flash') -> None:
        """
        Initialize MapHelper with map file path.

        Args:
            map_file_path: Path to the linker map file
            min_size: Minimum size threshold for filtering (bytes)
            sort_by: Sort key ('flash', 'ram', 'text', 'code', etc.)

        Raises:
            FileNotFoundError: If map file does not exist
        """
        self.map_file_path = map_file_path
        self.min_size = min_size
        self.sort_by = sort_by

        if not os.path.isfile(map_file_path):
            raise FileNotFoundError(f'Map file does not exist: {map_file_path}')

        # Regular expression patterns for parsing map file lines
        # Pattern: address(hex) size(hex) file_path(object_name)
        # Example: 0x00000000        0x8 armino/bk_rtos/libbk_rtos.a(rtos_pub.c.obj)
        self.size_pat = re.compile(r' +[^ ]* +0x([a-fA-F\d]+) +0x([a-fA-F\d]+) +(.*)\((.*)\)')

        # Parsed data structure: {file_path: {obj_name: {section: size}}}
        self.parsed_data: OrderedDict[str, OrderedDict[str, Dict[str, int]]] = OrderedDict()

        # Summary statistics
        self.summary: Dict[str, int] = {
            'total_flash': 0,
            'total_ram': 0,
            'total_text': 0,
            'total_code': 0,
            'total_rodata': 0,
            'total_data': 0,
            'total_bss': 0,
            'total_dec': 0,
            'library_count': 0,
            'object_count': 0,
        }

    def parse_flag(self, line: str) -> Optional[str]:
        """
        Parse section flag from line.

        Args:
            line: Line from map file

        Returns:
            Section type ('text', 'rodata', 'data', 'bss') or None
        """
        stripped_line = line.strip()
        for section_type, flags in self.SECTION_FLAGS.items():
            for flag in flags:
                if stripped_line.startswith(flag):
                    return section_type
        return None

    def add_parse_data(self, flag: str, size: int, file_path: str, obj_name: str) -> None:
        """
        Add parsed data to internal structure.

        Args:
            flag: Section type ('text', 'rodata', 'data', 'bss')
            size: Size in bytes
            file_path: Library file path
            obj_name: Object file name
        """
        if flag not in self.SECTION_FLAGS:
            raise ValueError(f'Invalid section flag: {flag}')

        if size < 0:
            raise ValueError(f'Invalid size: {size}')

        # Initialize data structure if needed
        if file_path not in self.parsed_data:
            self.parsed_data[file_path] = OrderedDict()

        if obj_name not in self.parsed_data[file_path]:
            self.parsed_data[file_path][obj_name] = {
                'rodata': 0,
                'text': 0,
                'data': 0,
                'bss': 0
            }

        self.parsed_data[file_path][obj_name][flag] += size

    def _calculate_sizes(self, code_size: int, rodata_size: int, data_size: int, bss_size: int) -> Tuple[int, int, int, int, str]:
        """
        Calculate derived sizes from base segment sizes.

        Args:
            code_size: Code segment size
            rodata_size: Read-only data segment size
            data_size: Initialized data segment size
            bss_size: Uninitialized data segment size

        Returns:
            Tuple of (flash_size, ram_size, text_size, dec_size, hex_size)
        """
        flash_size = code_size + rodata_size + data_size
        ram_size = data_size + bss_size
        text_size = code_size + rodata_size
        dec_size = text_size + data_size + bss_size
        hex_size = hex(dec_size)[2:]
        return flash_size, ram_size, text_size, dec_size, hex_size

    def start_parse(self) -> None:
        """
        Parse the map file and extract size information.

        Reads the map file line by line, identifies section types,
        and extracts size information for each object file.
        """
        current_section_type = None

        try:
            with open(self.map_file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in f:
                    # Check if line starts a new section
                    line_section_type = self.parse_flag(line)
                    if line_section_type is not None:
                        current_section_type = line_section_type
                        # For single-line entries (section name + addr/size/source
                        # on the same line), also try matching size_pat here instead
                        # of skipping to the next line.
                        size_match = self.size_pat.match(line)
                        if size_match:
                            address_hex = size_match.group(1)
                            size_hex = size_match.group(2)
                            file_path = size_match.group(3)
                            obj_name = size_match.group(4)

                            if address_hex != '0' and int(address_hex, 16) != 0:
                                size = int(size_hex, 16)
                                self.add_parse_data(current_section_type, size, file_path, obj_name)

                            current_section_type = None
                        continue

                    # Skip if we haven't found a section yet
                    if current_section_type is None:
                        continue

                    # Try to match size pattern
                    size_match = self.size_pat.match(line)
                    if size_match:
                        address_hex = size_match.group(1)
                        size_hex = size_match.group(2)
                        file_path = size_match.group(3)
                        obj_name = size_match.group(4)

                        # Skip if address is 0 (unused sections)
                        if address_hex != '0' and int(address_hex, 16) != 0:
                            size = int(size_hex, 16)
                            self.add_parse_data(current_section_type, size, file_path, obj_name)

                        # Clear section type after processing
                        current_section_type = None

        except IOError as e:
            raise IOError(f'Error reading map file: {e}') from e

    def _calculate_summary(self) -> None:
        """
        Calculate summary statistics from parsed data.

        Computes totals for:
        - Flash usage (code + rodata + data)
        - RAM usage (data + bss)
        - Text segment (code + rodata)
        - Individual segments (code, rodata, data, bss)
        - Library and object counts
        """
        self.summary = {
            'total_flash': 0,
            'total_ram': 0,
            'total_text': 0,
            'total_code': 0,
            'total_rodata': 0,
            'total_data': 0,
            'total_bss': 0,
            'total_dec': 0,
            'library_count': len(self.parsed_data),
            'object_count': 0,
        }

        for library_file in self.parsed_data.values():
            for obj_data in library_file.values():
                self.summary['total_code'] += obj_data['text']
                self.summary['total_rodata'] += obj_data['rodata']
                self.summary['total_data'] += obj_data['data']
                self.summary['total_bss'] += obj_data['bss']
                self.summary['object_count'] += 1

        self.summary['total_text'] = self.summary['total_code'] + self.summary['total_rodata']
        self.summary['total_flash'] = self.summary['total_code'] + self.summary['total_rodata'] + self.summary['total_data']
        self.summary['total_ram'] = self.summary['total_data'] + self.summary['total_bss']
        self.summary['total_dec'] = self.summary['total_text'] + self.summary['total_data'] + self.summary['total_bss']

    def _get_sorted_libraries(self) -> List[Tuple[str, Dict[str, int]]]:
        """
        Get libraries sorted by specified key with minimum size filtering applied.

        This method:
        - Calculates totals for each library (flash, ram, text, code, rodata, data, bss)
        - Applies minimum size filter (min_size) based on sort_by field
        - Sorts libraries in descending order by sort_by field

        Returns:
            List of (library_file, totals) tuples sorted by sort_by key in descending order.
            Only libraries meeting the minimum size threshold are included.
        """
        libraries = []
        for library_file in self.parsed_data.keys():
            totals = {
                'text': 0,
                'rodata': 0,
                'data': 0,
                'bss': 0,
            }
            for obj_data in self.parsed_data[library_file].values():
                totals['text'] += obj_data['text']
                totals['rodata'] += obj_data['rodata']
                totals['data'] += obj_data['data']
                totals['bss'] += obj_data['bss']

            code_size = totals['text']
            flash_size, ram_size, text_size, dec_size, _ = self._calculate_sizes(
                code_size, totals['rodata'], totals['data'], totals['bss']
            )

            totals['flash'] = flash_size
            totals['ram'] = ram_size
            totals['text'] = text_size
            totals['code'] = code_size
            totals['dec'] = dec_size

            # Apply minimum size filter
            if totals.get(self.sort_by, 0) >= self.min_size:
                libraries.append((library_file, totals))

        # Sort by specified key (descending)
        libraries.sort(key=lambda x: x[1].get(self.sort_by, 0), reverse=True)
        return libraries

    def format_title_line(self) -> str:
        """
        Format title line for text output.

        Returns:
            Formatted title line string
        """
        columns = ['flash', 'ram', 'text', 'code', 'rodata', 'data', 'bss', 'dec', 'hex', 'filename']
        return '\t'.join(col.rjust(self.COLUMN_WIDTH) for col in columns) + '\n'

    def format_size_line(self, code_size: int, rodata_size: int, data_size: int, bss_size: int,
                        filename: str, obj_name: str, library_path: str = '', percentage: float = 0.0) -> Tuple[List, str]:
        """
        Format size line for both CSV and text output.

        Args:
            code_size: Code segment size
            rodata_size: Read-only data segment size
            data_size: Initialized data segment size
            bss_size: Uninitialized data segment size
            filename: Library file path
            obj_name: Object file name
            percentage: Percentage of total (optional)

        Returns:
            Tuple of (csv_row_list, formatted_text_line)
        """
        flash_size, ram_size, text_size, dec_size, hex_size = self._calculate_sizes(
            code_size, rodata_size, data_size, bss_size
        )

        lib_basename = os.path.basename(filename)

        # CSV row data
        csv_row = [
            flash_size, ram_size, text_size, code_size, rodata_size,
            data_size, bss_size, dec_size, f'`{hex_size}',
            f'{obj_name} (ex {lib_basename})', lib_basename, library_path
        ]

        # Formatted text line
        text_line = '\t'.join([
            str(flash_size).rjust(self.COLUMN_WIDTH),
            str(ram_size).rjust(self.COLUMN_WIDTH),
            str(text_size).rjust(self.COLUMN_WIDTH),
            str(code_size).rjust(self.COLUMN_WIDTH),
            str(rodata_size).rjust(self.COLUMN_WIDTH),
            str(data_size).rjust(self.COLUMN_WIDTH),
            str(bss_size).rjust(self.COLUMN_WIDTH),
            str(dec_size).rjust(self.COLUMN_WIDTH),
            hex_size.rjust(self.COLUMN_WIDTH),
            f'{obj_name} (ex {lib_basename})'
        ]) + '\n'

        return csv_row, text_line

    def format_total_size_line(self, code_size: int, rodata_size: int, data_size: int, bss_size: int,
                               filename: str, library_path: str = '', percentage: float = 0.0) -> Tuple[List, str]:
        """
        Format total size line for library file.

        Args:
            code_size: Total code segment size
            rodata_size: Total read-only data segment size
            data_size: Total initialized data segment size
            bss_size: Total uninitialized data segment size
            filename: Library file path
            percentage: Percentage of total (optional)

        Returns:
            Tuple of (csv_row_list, formatted_text_line)
        """
        flash_size, ram_size, text_size, dec_size, hex_size = self._calculate_sizes(
            code_size, rodata_size, data_size, bss_size
        )

        lib_basename = os.path.basename(filename)

        # CSV row data
        csv_row = [
            flash_size, ram_size, text_size, code_size, rodata_size,
            data_size, bss_size, dec_size, f'`{hex_size}', lib_basename, library_path
        ]

        # Formatted text line
        text_line = '\t'.join([
            str(flash_size).rjust(self.COLUMN_WIDTH),
            str(ram_size).rjust(self.COLUMN_WIDTH),
            str(text_size).rjust(self.COLUMN_WIDTH),
            str(code_size).rjust(self.COLUMN_WIDTH),
            str(rodata_size).rjust(self.COLUMN_WIDTH),
            str(data_size).rjust(self.COLUMN_WIDTH),
            str(bss_size).rjust(self.COLUMN_WIDTH),
            str(dec_size).rjust(self.COLUMN_WIDTH),
            hex_size.rjust(self.COLUMN_WIDTH),
            lib_basename
        ]) + '\n'

        return csv_row, text_line

    def _get_output_basename(self) -> str:
        """
        Generate output file basename based on input map file name.

        Returns:
            Base name for output files (without extension)
        """
        map_basename = os.path.basename(self.map_file_path)
        # Remove .map extension if present
        if map_basename.endswith('.map'):
            base_name = map_basename[:-4]
        else:
            base_name = map_basename

        # Generate output basename: <input_name>_size_map
        return f'{base_name}_size_map'

    def _get_top_objects(self, top_n: int = 15) -> List[Dict]:
        """
        Get top N objects sorted by Flash usage with detailed segment breakdown.

        This method collects all objects from all libraries, calculates their Flash size
        (code + rodata + data), and returns the top N objects sorted in descending order.
        Used for generating Top15 visualization sheet in Excel output.

        Args:
            top_n: Number of top objects to return (default: 15)

        Returns:
            List of dictionaries sorted by flash size (descending), each containing:
            {
                'library': str,      # Library file path
                'object': str,       # Object file name
                'flash': int,        # Total Flash size (code + rodata + data)
                'text': int,         # Code segment size
                'rodata': int,       # Read-only data segment size
                'data': int,         # Initialized data segment size
                'bss': int,          # Uninitialized data segment size
                'ram': int,          # Total RAM size (data + bss)
                'flash_pct': float   # Flash size as percentage of total Flash
            }
        """
        top_items = []
        total_flash = self.summary.get('total_flash', 0)

        for library_file, objs in self.parsed_data.items():
            for obj_name, obj_data in objs.items():
                flash_size, ram_size, text_size, _, _ = self._calculate_sizes(
                    obj_data.get('text', 0),
                    obj_data.get('rodata', 0),
                    obj_data.get('data', 0),
                    obj_data.get('bss', 0)
                )
                flash_pct = (flash_size / total_flash * 100) if total_flash > 0 else 0.0
                top_items.append({
                    'library': library_file,
                    'object': obj_name,
                    'flash': flash_size,
                    'text': obj_data.get('text', 0),
                    'rodata': obj_data.get('rodata', 0),
                    'data': obj_data.get('data', 0),
                    'bss': obj_data.get('bss', 0),
                    'ram': ram_size,
                    'flash_pct': flash_pct,
                })

        # Sort and take top N
        top_items.sort(key=lambda x: x['flash'], reverse=True)
        return top_items[:top_n]

    def _get_top_libraries_by_field(self, field: str, top_n: int = 15) -> List[Dict]:
        """
        Get top N libraries sorted by specified field with detailed segment breakdown.

        This method collects all libraries, sorts them by the specified field (text, bss, data, ram),
        and returns the top N libraries with their segment breakdown. Used for generating
        library-level Top15 visualization sheets in Excel output.

        Args:
            field: Field to sort by ('text', 'bss', 'data', 'ram')
            top_n: Number of top libraries to return (default: 15)

        Returns:
            List of dictionaries sorted by specified field (descending), each containing:
            {
                'library': str,      # Library file path
                'text': int,         # Text segment size (code + rodata)
                'code': int,         # Code segment size (for text breakdown)
                'rodata': int,       # Read-only data segment size (for text breakdown)
                'data': int,         # Initialized data segment size
                'bss': int,          # Uninitialized data segment size
                'ram': int,          # Total RAM size (data + bss)
                'flash': int,        # Total Flash size (code + rodata + data)
                'field_value': int,  # Value of the sorting field
                'field_pct': float   # Field value as percentage of total
            }
        """
        if field not in ['text', 'bss', 'data', 'ram']:
            raise ValueError(f'Invalid field: {field}. Must be one of: text, bss, data, ram')

        top_libs = []
        total_value = self.summary.get(f'total_{field}', 0)
        if field == 'text':
            total_value = self.summary.get('total_text', 0)
        elif field == 'ram':
            total_value = self.summary.get('total_ram', 0)

        for library_file in self.parsed_data.keys():
            totals = {
                'text': 0,
                'rodata': 0,
                'data': 0,
                'bss': 0,
            }

            # Aggregate all objects in the library
            for obj_data in self.parsed_data[library_file].values():
                totals['text'] += obj_data.get('text', 0)
                totals['rodata'] += obj_data.get('rodata', 0)
                totals['data'] += obj_data.get('data', 0)
                totals['bss'] += obj_data.get('bss', 0)

            # Calculate derived values
            flash_size, ram_size, text_size, _, _ = self._calculate_sizes(
                totals['text'], totals['rodata'], totals['data'], totals['bss']
            )

            # Get the field value
            if field == 'text':
                field_value = text_size
            elif field == 'bss':
                field_value = totals['bss']
            elif field == 'data':
                field_value = totals['data']
            elif field == 'ram':
                field_value = ram_size
            else:
                field_value = 0

            field_pct = (field_value / total_value * 100) if total_value > 0 else 0.0

            top_libs.append({
                'library': library_file,
                'text': text_size,
                'code': totals['text'],  # code segment
                'rodata': totals['rodata'],
                'data': totals['data'],
                'bss': totals['bss'],
                'ram': ram_size,
                'flash': flash_size,
                'field_value': field_value,
                'field_pct': field_pct,
            })

        # Sort by field value and take top N
        top_libs.sort(key=lambda x: x['field_value'], reverse=True)
        return top_libs[:top_n]

    def _save_to_excel(self, output_dir: str, output_basename: str, sorted_libraries: List[Tuple[str, Dict[str, int]]]) -> None:
        """
        Save data to Excel file with analysis sheets (Total, Detail, and Top15 Libraries by field).

        Args:
            output_dir: Output directory path
            output_basename: Base name for output file
            sorted_libraries: Sorted list of libraries with totals
        """
        if not EXCEL_SUPPORT:
            print('Warning: openpyxl not available, skipping Excel output. Install with: pip install openpyxl', file=sys.stderr)
            return

        excel_path = os.path.join(output_dir, f'{output_basename}.xlsx')

        try:
            wb = Workbook()

            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            ws_total = wb.create_sheet('Total', 0)
            ws_total.append(self.CSV_COLUMNS)

            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws_total[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            ws_detail = wb.create_sheet('Detail', 1)
            ws_detail.append(self.CSV_DETAIL_COLUMNS)

            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            total_rows = []
            detail_rows = []

            for library_file, library_totals in sorted_libraries:
                lib_path = os.path.dirname(library_file)
                flash_size = library_totals['flash']
                percentage = (flash_size / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0

                csv_total_row, _ = self.format_total_size_line(
                    library_totals['code'],
                    library_totals['rodata'],
                    library_totals['data'],
                    library_totals['bss'],
                    library_file,
                    library_path=lib_path,
                    percentage=percentage
                )

                ws_total.append(csv_total_row)
                total_rows.append({
                    'row': list(csv_total_row),
                    'flash': flash_size,
                    'library': library_file
                })

                for obj_file, obj_data in self.parsed_data[library_file].items():
                    obj_flash, _, _, _, _ = self._calculate_sizes(
                        obj_data['text'], obj_data['rodata'], obj_data['data'], obj_data['bss']
                    )
                    obj_percentage = (obj_flash / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0

                    csv_detail_row, _ = self.format_size_line(
                        obj_data['text'],
                        obj_data['rodata'],
                        obj_data['data'],
                        obj_data['bss'],
                        library_file,
                        obj_file,
                        library_path=lib_path,
                        percentage=obj_percentage
                    )

                    ws_detail.append(csv_detail_row)
                    detail_rows.append({
                        'row': list(csv_detail_row),
                        'flash': obj_flash,
                        'library': library_file,
                        'obj': obj_file
                    })

                ws_detail.append(csv_total_row)
                detail_rows.append({
                    'row': list(csv_total_row),
                    'flash': flash_size,
                    'library': library_file,
                    'obj': library_file
                })

            # Create library-level Top15 sheets for different fields
            # Top15_Text: Top 15 libraries by text size
            top_libs_text = self._get_top_libraries_by_field('text', top_n=15)
            ws_top_text = wb.create_sheet('Top15_Text', 2)
            ws_top_text.append(['rank', 'library', 'text', 'code', 'rodata', 'text_pct'])
            for cell in ws_top_text[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, lib in enumerate(top_libs_text, 1):
                ws_top_text.append([
                    idx,
                    lib['library'],
                    lib['text'],
                    lib['code'],
                    lib['rodata'],
                    f"{lib['field_pct']:.2f}%"
                ])

            if top_libs_text:
                chart_text = BarChart()
                chart_text.type = "col"
                chart_text.grouping = "stacked"
                chart_text.title = "Top15 Libraries by Text (stacked: code/rodata)"
                chart_text.y_axis.title = "Size (Bytes)"
                chart_text.x_axis.title = "Library"
                data_text = Reference(ws_top_text, min_col=4, max_col=5, min_row=1, max_row=len(top_libs_text) + 1)
                cats_text = Reference(ws_top_text, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_text) + 1)
                chart_text.add_data(data_text, titles_from_data=True)
                chart_text.set_categories(cats_text)
                chart_text.height = 15
                chart_text.width = 30
                ws_top_text.add_chart(chart_text, "G2")

            # Top15_BSS: Top 15 libraries by BSS size
            top_libs_bss = self._get_top_libraries_by_field('bss', top_n=15)
            ws_top_bss = wb.create_sheet('Top15_BSS', 3)
            ws_top_bss.append(['rank', 'library', 'bss', 'bss_pct'])
            for cell in ws_top_bss[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, lib in enumerate(top_libs_bss, 1):
                ws_top_bss.append([
                    idx,
                    lib['library'],
                    lib['bss'],
                    f"{lib['field_pct']:.2f}%"
                ])

            if top_libs_bss:
                chart_bss = BarChart()
                chart_bss.type = "col"
                chart_bss.grouping = "standard"
                chart_bss.title = "Top15 Libraries by BSS"
                chart_bss.y_axis.title = "Size (Bytes)"
                chart_bss.x_axis.title = "Library"
                data_bss = Reference(ws_top_bss, min_col=3, max_col=3, min_row=1, max_row=len(top_libs_bss) + 1)
                cats_bss = Reference(ws_top_bss, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_bss) + 1)
                chart_bss.add_data(data_bss, titles_from_data=True)
                chart_bss.set_categories(cats_bss)
                chart_bss.height = 15
                chart_bss.width = 30
                ws_top_bss.add_chart(chart_bss, "E2")

            # Top15_Data: Top 15 libraries by Data (rwdata) size
            top_libs_data = self._get_top_libraries_by_field('data', top_n=15)
            ws_top_data = wb.create_sheet('Top15_Data', 4)
            ws_top_data.append(['rank', 'library', 'data', 'data_pct'])
            for cell in ws_top_data[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, lib in enumerate(top_libs_data, 1):
                ws_top_data.append([
                    idx,
                    lib['library'],
                    lib['data'],
                    f"{lib['field_pct']:.2f}%"
                ])

            if top_libs_data:
                chart_data = BarChart()
                chart_data.type = "col"
                chart_data.grouping = "standard"
                chart_data.title = "Top15 Libraries by Data (rwdata)"
                chart_data.y_axis.title = "Size (Bytes)"
                chart_data.x_axis.title = "Library"
                data_data = Reference(ws_top_data, min_col=3, max_col=3, min_row=1, max_row=len(top_libs_data) + 1)
                cats_data = Reference(ws_top_data, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_data) + 1)
                chart_data.add_data(data_data, titles_from_data=True)
                chart_data.set_categories(cats_data)
                chart_data.height = 15
                chart_data.width = 30
                ws_top_data.add_chart(chart_data, "E2")

            # Top15_RAM: Top 15 libraries by RAM size
            top_libs_ram = self._get_top_libraries_by_field('ram', top_n=15)
            ws_top_ram = wb.create_sheet('Top15_RAM', 5)
            ws_top_ram.append(['rank', 'library', 'ram', 'data', 'bss', 'ram_pct'])
            for cell in ws_top_ram[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, lib in enumerate(top_libs_ram, 1):
                ws_top_ram.append([
                    idx,
                    lib['library'],
                    lib['ram'],
                    lib['data'],
                    lib['bss'],
                    f"{lib['field_pct']:.2f}%"
                ])

            if top_libs_ram:
                chart_ram = BarChart()
                chart_ram.type = "col"
                chart_ram.grouping = "stacked"
                chart_ram.title = "Top15 Libraries by RAM (stacked: data/bss)"
                chart_ram.y_axis.title = "Size (Bytes)"
                chart_ram.x_axis.title = "Library"
                data_ram = Reference(ws_top_ram, min_col=4, max_col=5, min_row=1, max_row=len(top_libs_ram) + 1)
                cats_ram = Reference(ws_top_ram, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_ram) + 1)
                chart_ram.add_data(data_ram, titles_from_data=True)
                chart_ram.set_categories(cats_ram)
                chart_ram.height = 15
                chart_ram.width = 30
                ws_top_ram.add_chart(chart_ram, "G2")

            def prepare_sheet(sheet_name: str, headers: List[str]):
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                return ws

            total_headers_with_rank = ['rank'] + self.CSV_COLUMNS
            ws_total_flash = prepare_sheet('SortedByFlashSize', total_headers_with_rank)
            total_flash_sorted = sorted(total_rows, key=lambda item: item['flash'], reverse=True)
            for idx, item in enumerate(total_flash_sorted, start=1):
                ws_total_flash.append([idx] + item['row'])
            self._auto_adjust_columns(ws_total_flash)
            ws_total_flash.freeze_panes = 'A2'

            ws_total_by_library = prepare_sheet('SortedByLibraryName', total_headers_with_rank)
            total_by_library_sorted = sorted(total_rows, key=lambda item: os.path.basename(item['library']).lower())
            for idx, item in enumerate(total_by_library_sorted, start=1):
                ws_total_by_library.append([idx] + item['row'])
            self._auto_adjust_columns(ws_total_by_library)
            ws_total_by_library.freeze_panes = 'A2'

            detail_headers_with_rank = ['rank'] + self.CSV_DETAIL_COLUMNS
            ws_detail_flash = prepare_sheet('SortedByFlashSizeDetail', detail_headers_with_rank)
            detail_flash_sorted = sorted(detail_rows, key=lambda item: item['flash'], reverse=True)
            for idx, item in enumerate(detail_flash_sorted, start=1):
                ws_detail_flash.append([idx] + item['row'])
            self._auto_adjust_columns(ws_detail_flash)
            ws_detail_flash.freeze_panes = 'A2'

            ws_detail_by_library = prepare_sheet('SortedByLibraryNameDetail', detail_headers_with_rank)
            detail_by_library_sorted = sorted(
                detail_rows,
                key=lambda item: (os.path.basename(item['library']).lower(), item['obj'])
            )
            for idx, item in enumerate(detail_by_library_sorted, start=1):
                ws_detail_by_library.append([idx] + item['row'])
            self._auto_adjust_columns(ws_detail_by_library)
            ws_detail_by_library.freeze_panes = 'A2'

            # Auto-adjust column widths and freeze panes for main sheets
            for ws in [ws_total, ws_detail, ws_top_text, ws_top_bss, ws_top_data, ws_top_ram]:
                self._auto_adjust_columns(ws)
                ws.freeze_panes = 'A2'

            wb.save(excel_path)

        except Exception as e:
            raise IOError(f'Error writing Excel file: {e}') from e

    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths for the given worksheet."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


    @staticmethod
    def _generate_merged_excel(all_map_helpers: List[Tuple[str, 'MapHelper']],
                               processed_files: List[Tuple[str, str, str]],
                               args) -> None:
        """
        Generate a merged Excel file containing all map files' data in separate sheets.

        Args:
            all_map_helpers: List of (map_file_path, MapHelper instance) tuples
            processed_files: List of (map_file, output_dir, output_basename) tuples
            args: Parsed command-line arguments
        """
        if not EXCEL_SUPPORT:
            return

        # Determine output directory (use first file's directory or current directory)
        if processed_files:
            output_dir = processed_files[0][1]
        else:
            output_dir = os.getcwd()

        # Generate merged filename
        merged_filename = 'merged_size_map.xlsx'
        merged_path = os.path.join(output_dir, merged_filename)

        try:
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for map_file, mh in all_map_helpers:
                # Generate sheet name from map file basename
                map_basename = os.path.basename(map_file)
                if map_basename.endswith('.map'):
                    sheet_base = map_basename[:-4]
                else:
                    sheet_base = map_basename

                # Limit sheet name length (Excel limit is 31 characters)
                sheet_base = sheet_base[:25] if len(sheet_base) > 25 else sheet_base

                sorted_libraries = mh._get_sorted_libraries()

                # Create "Total" sheet for this map file
                sheet_total_name = f'{sheet_base}_Total'
                ws_total = wb.create_sheet(sheet_total_name)
                ws_total.append(mh.CSV_COLUMNS)

                # Style header row
                for cell in ws_total[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Write total data
                for library_file, library_totals in sorted_libraries:
                    flash_size = library_totals['flash']
                    percentage = (flash_size / mh.summary['total_flash'] * 100) if mh.summary['total_flash'] > 0 else 0.0

                    lib_path = os.path.dirname(library_file)
                    csv_total_row, _ = mh.format_total_size_line(
                        library_totals['code'],
                        library_totals['rodata'],
                        library_totals['data'],
                        library_totals['bss'],
                        library_file,
                        library_path=lib_path,
                        percentage=percentage
                    )
                    ws_total.append(csv_total_row)

                # Auto-adjust column widths
                for column in ws_total.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_total.column_dimensions[column_letter].width = adjusted_width
                ws_total.freeze_panes = 'A2'

                # Create "Detail" sheet for this map file
                sheet_detail_name = f'{sheet_base}_Detail'
                ws_detail = wb.create_sheet(sheet_detail_name)
                ws_detail.append(mh.CSV_DETAIL_COLUMNS)

                # Style header row
                for cell in ws_detail[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Write detail data
                for library_file, library_totals in sorted_libraries:
                    lib_path = os.path.dirname(library_file)
                    for obj_file in mh.parsed_data[library_file].keys():
                        obj_data = mh.parsed_data[library_file][obj_file]

                        flash_size, _, _, _, _ = mh._calculate_sizes(
                            obj_data['text'], obj_data['rodata'], obj_data['data'], obj_data['bss']
                        )
                        percentage = (flash_size / mh.summary['total_flash'] * 100) if mh.summary['total_flash'] > 0 else 0.0

                        csv_detail_row, _ = mh.format_size_line(
                            obj_data['text'],
                            obj_data['rodata'],
                            obj_data['data'],
                            obj_data['bss'],
                            library_file,
                            obj_file,
                            library_path=lib_path,
                            percentage=percentage
                        )
                        ws_detail.append(csv_detail_row)

                    # Also add library total to detail sheet
                    flash_size = library_totals['flash']
                    percentage = (flash_size / mh.summary['total_flash'] * 100) if mh.summary['total_flash'] > 0 else 0.0

                    csv_total_row, _ = mh.format_total_size_line(
                        library_totals['code'],
                        library_totals['rodata'],
                        library_totals['data'],
                        library_totals['bss'],
                        library_file,
                        library_path=lib_path,
                        percentage=percentage
                    )
                    ws_detail.append(csv_total_row)

                # Auto-adjust column widths
                for column in ws_detail.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_detail.column_dimensions[column_letter].width = adjusted_width
                ws_detail.freeze_panes = 'A2'

                # Create library-level Top15 sheets for this map file
                # Top15_Text
                top_libs_text = mh._get_top_libraries_by_field('text', top_n=15)
                sheet_top_text_name = f'{sheet_base}_Top15_Text'
                ws_top_text = wb.create_sheet(sheet_top_text_name)
                ws_top_text.append(['rank', 'library', 'text', 'code', 'rodata', 'text_pct'])
                for cell in ws_top_text[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for idx, lib in enumerate(top_libs_text, 1):
                    ws_top_text.append([
                        idx, lib['library'], lib['text'], lib['code'], lib['rodata'],
                        f"{lib['field_pct']:.2f}%"
                    ])

                if top_libs_text:
                    chart_text = BarChart()
                    chart_text.type = "col"
                    chart_text.grouping = "stacked"
                    chart_text.title = f"Top15 Libraries by Text - {sheet_base}"
                    chart_text.y_axis.title = "Size (Bytes)"
                    chart_text.x_axis.title = "Library"
                    data_text = Reference(ws_top_text, min_col=4, max_col=5, min_row=1, max_row=len(top_libs_text) + 1)
                    cats_text = Reference(ws_top_text, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_text) + 1)
                    chart_text.add_data(data_text, titles_from_data=True)
                    chart_text.set_categories(cats_text)
                    chart_text.height = 15
                    chart_text.width = 30
                    ws_top_text.add_chart(chart_text, "G2")

                # Top15_BSS
                top_libs_bss = mh._get_top_libraries_by_field('bss', top_n=15)
                sheet_top_bss_name = f'{sheet_base}_Top15_BSS'
                ws_top_bss = wb.create_sheet(sheet_top_bss_name)
                ws_top_bss.append(['rank', 'library', 'bss', 'bss_pct'])
                for cell in ws_top_bss[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for idx, lib in enumerate(top_libs_bss, 1):
                    ws_top_bss.append([
                        idx, lib['library'], lib['bss'], f"{lib['field_pct']:.2f}%"
                    ])

                if top_libs_bss:
                    chart_bss = BarChart()
                    chart_bss.type = "col"
                    chart_bss.grouping = "standard"
                    chart_bss.title = f"Top15 Libraries by BSS - {sheet_base}"
                    chart_bss.y_axis.title = "Size (Bytes)"
                    chart_bss.x_axis.title = "Library"
                    data_bss = Reference(ws_top_bss, min_col=3, max_col=3, min_row=1, max_row=len(top_libs_bss) + 1)
                    cats_bss = Reference(ws_top_bss, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_bss) + 1)
                    chart_bss.add_data(data_bss, titles_from_data=True)
                    chart_bss.set_categories(cats_bss)
                    chart_bss.height = 15
                    chart_bss.width = 30
                    ws_top_bss.add_chart(chart_bss, "E2")

                # Top15_Data
                top_libs_data = mh._get_top_libraries_by_field('data', top_n=15)
                sheet_top_data_name = f'{sheet_base}_Top15_Data'
                ws_top_data = wb.create_sheet(sheet_top_data_name)
                ws_top_data.append(['rank', 'library', 'data', 'data_pct'])
                for cell in ws_top_data[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for idx, lib in enumerate(top_libs_data, 1):
                    ws_top_data.append([
                        idx, lib['library'], lib['data'], f"{lib['field_pct']:.2f}%"
                    ])

                if top_libs_data:
                    chart_data = BarChart()
                    chart_data.type = "col"
                    chart_data.grouping = "standard"
                    chart_data.title = f"Top15 Libraries by Data - {sheet_base}"
                    chart_data.y_axis.title = "Size (Bytes)"
                    chart_data.x_axis.title = "Library"
                    data_data = Reference(ws_top_data, min_col=3, max_col=3, min_row=1, max_row=len(top_libs_data) + 1)
                    cats_data = Reference(ws_top_data, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_data) + 1)
                    chart_data.add_data(data_data, titles_from_data=True)
                    chart_data.set_categories(cats_data)
                    chart_data.height = 15
                    chart_data.width = 30
                    ws_top_data.add_chart(chart_data, "E2")

                # Top15_RAM
                top_libs_ram = mh._get_top_libraries_by_field('ram', top_n=15)
                sheet_top_ram_name = f'{sheet_base}_Top15_RAM'
                ws_top_ram = wb.create_sheet(sheet_top_ram_name)
                ws_top_ram.append(['rank', 'library', 'ram', 'data', 'bss', 'ram_pct'])
                for cell in ws_top_ram[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                for idx, lib in enumerate(top_libs_ram, 1):
                    ws_top_ram.append([
                        idx, lib['library'], lib['ram'], lib['data'], lib['bss'],
                        f"{lib['field_pct']:.2f}%"
                    ])

                if top_libs_ram:
                    chart_ram = BarChart()
                    chart_ram.type = "col"
                    chart_ram.grouping = "stacked"
                    chart_ram.title = f"Top15 Libraries by RAM - {sheet_base}"
                    chart_ram.y_axis.title = "Size (Bytes)"
                    chart_ram.x_axis.title = "Library"
                    data_ram = Reference(ws_top_ram, min_col=4, max_col=5, min_row=1, max_row=len(top_libs_ram) + 1)
                    cats_ram = Reference(ws_top_ram, min_col=2, max_col=2, min_row=2, max_row=len(top_libs_ram) + 1)
                    chart_ram.add_data(data_ram, titles_from_data=True)
                    chart_ram.set_categories(cats_ram)
                    chart_ram.height = 15
                    chart_ram.width = 30
                    ws_top_ram.add_chart(chart_ram, "G2")

                # Auto-adjust column widths for library Top15 sheets
                for ws_lib_top in [ws_top_text, ws_top_bss, ws_top_data, ws_top_ram]:
                    for column in ws_lib_top.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws_lib_top.column_dimensions[column_letter].width = adjusted_width
                    ws_lib_top.freeze_panes = 'A2'

            wb.save(merged_path)
            print(f'\n{"=" * 80}')
            print(f'Successfully generated merged Excel file: {merged_path}')
            print(f'  Contains {len(all_map_helpers) * 6} sheets (Total, Detail, and 4 Top15 Library sheets for each map file)')
            print(f'  Each map file has separate sheets - data is NOT merged together')
            sheet_names = []
            for map_file, _ in all_map_helpers:
                map_basename = os.path.basename(map_file)
                if map_basename.endswith('.map'):
                    sheet_base = map_basename[:-4]
                else:
                    sheet_base = map_basename
                sheet_base = sheet_base[:25] if len(sheet_base) > 25 else sheet_base
                sheet_names.append(f'{sheet_base}_Total, {sheet_base}_Detail, {sheet_base}_Top15_Text, {sheet_base}_Top15_BSS, {sheet_base}_Top15_Data, {sheet_base}_Top15_RAM')
            print(f'  Sheet names: {", ".join(sheet_names)}')
            print('=' * 80)

        except Exception as e:
            raise IOError(f'Error writing merged Excel file: {e}') from e

    def save_to_file(self, output_format: str = 'all', excel_format: bool = True) -> None:
        """
        Save parsed data to output files.

        Args:
            output_format: Output format ('all', 'txt', 'csv', 'json', 'excel')
            excel_format: Whether to generate Excel file with multiple sheets (Total, Detail, Top15_Text, Top15_BSS, Top15_Data, Top15_RAM) (default: True)

        Generates output files based on input map file name:
        - <input_name>_size_map.txt: Human-readable text format
        - <input_name>_size_map_total.csv: Summary statistics per library (when output_format includes 'csv' or 'all')
        - <input_name>_size_map_detail.csv: Detailed statistics per object file (when output_format includes 'csv' or 'all')
        - <input_name>_size_map.xlsx: Excel file with multiple sheets:
          * "Total" sheet: Summary statistics per library file (.a files)
          * "Detail" sheet: Detailed statistics per object file (.obj files)
          * "Top15_Text" sheet: Top 15 libraries by text size (code+rodata) with stacked chart
          * "Top15_BSS" sheet: Top 15 libraries by BSS size with bar chart
          * "Top15_Data" sheet: Top 15 libraries by Data (rwdata) size with bar chart
          * "Top15_RAM" sheet: Top 15 libraries by RAM size (data+bss) with stacked chart
          (when excel_format=True and output_format includes 'excel', 'csv', or 'all')
        - <input_name>_size_map.json: JSON format (when output_format includes 'json' or 'all')
        """
        # Calculate summary statistics
        self._calculate_summary()

        output_dir = os.path.dirname(self.map_file_path)
        output_basename = self._get_output_basename()

        # Generate sorted libraries
        sorted_libraries = self._get_sorted_libraries()

        # Save Excel format (two sheets in one file)
        if excel_format and output_format in ('all', 'excel', 'csv'):
            self._save_to_excel(output_dir, output_basename, sorted_libraries)

        # Save text and CSV formats (separate files)
        if output_format in ('all', 'txt', 'csv'):
            txt_path = os.path.join(output_dir, f'{output_basename}.txt')
            csv_total_path = os.path.join(output_dir, f'{output_basename}_total.csv')
            csv_detail_path = os.path.join(output_dir, f'{output_basename}_detail.csv')

            try:
                with open(txt_path, 'w', encoding='utf-8') as f_txt, \
                     open(csv_total_path, 'w', newline='', encoding='utf-8') as f_csv_total, \
                     open(csv_detail_path, 'w', newline='', encoding='utf-8') as f_csv_detail:

                    # Initialize CSV writers
                    csv_total_writer = csv.writer(f_csv_total)
                    csv_detail_writer = csv.writer(f_csv_detail)

                    # Write CSV headers
                    csv_total_writer.writerow(self.CSV_COLUMNS)
                    csv_detail_writer.writerow(self.CSV_DETAIL_COLUMNS)

                    # Write summary header to text file
                    f_txt.write('=' * 80 + '\n')
                    f_txt.write('Size Map Summary\n')
                    f_txt.write('=' * 80 + '\n')
                    f_txt.write(f'Total Flash: {self.summary["total_flash"]:,} bytes ({self.summary["total_flash"]/1024:.2f} KB)\n')
                    f_txt.write(f'Total RAM:   {self.summary["total_ram"]:,} bytes ({self.summary["total_ram"]/1024:.2f} KB)\n')
                    f_txt.write(f'Total Code:  {self.summary["total_code"]:,} bytes ({self.summary["total_code"]/1024:.2f} KB)\n')
                    f_txt.write(f'Total Data:  {self.summary["total_data"]:,} bytes ({self.summary["total_data"]/1024:.2f} KB)\n')
                    f_txt.write(f'Libraries:   {self.summary["library_count"]}\n')
                    f_txt.write(f'Objects:    {self.summary["object_count"]}\n')
                    f_txt.write('=' * 80 + '\n\n')

                    # Process each library file (sorted)
                    for library_file, library_totals in sorted_libraries:
                        # Write title line for text output
                        f_txt.write(self.format_title_line())

                        # Process each object file in the library
                        lib_path = os.path.dirname(library_file)
                        for obj_file in self.parsed_data[library_file].keys():
                            obj_data = self.parsed_data[library_file][obj_file]

                            # Calculate percentage
                            flash_size, _, _, _, _ = self._calculate_sizes(
                                obj_data['text'], obj_data['rodata'], obj_data['data'], obj_data['bss']
                            )
                            percentage = (flash_size / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0

                            # Format and write detail line
                            csv_detail_row, text_detail_line = self.format_size_line(
                                obj_data['text'],
                                obj_data['rodata'],
                                obj_data['data'],
                                obj_data['bss'],
                                library_file,
                                obj_file,
                                library_path=lib_path,
                                percentage=percentage
                            )
                            f_txt.write(text_detail_line)
                            csv_detail_writer.writerow(csv_detail_row)

                        # Calculate percentage for library total
                        flash_size = library_totals['flash']
                        percentage = (flash_size / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0

                        # Format and write total line for library
                        csv_total_row, text_total_line = self.format_total_size_line(
                            library_totals['code'],
                            library_totals['rodata'],
                            library_totals['data'],
                            library_totals['bss'],
                            library_file,
                            library_path=lib_path,
                            percentage=percentage
                        )
                        f_txt.write(text_total_line)
                        f_txt.write('\n')

                        # Write total to CSV files
                        csv_total_writer.writerow(csv_total_row)

                        # Also write total to detail CSV (for consistency)
                        csv_detail_writer.writerow(csv_total_row)

            except IOError as e:
                raise IOError(f'Error writing output files: {e}') from e

        # Save JSON format
        if output_format in ('all', 'json'):
            json_path = os.path.join(output_dir, f'{output_basename}.json')
            try:
                json_data = {
                    'summary': self.summary,
                    'libraries': []
                }

                for library_file, library_totals in sorted_libraries:
                    library_info = {
                        'filename': library_file,
                        'sizes': library_totals,
                        'percentage': {
                            'flash': (library_totals['flash'] / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0,
                            'ram': (library_totals['ram'] / self.summary['total_ram'] * 100) if self.summary['total_ram'] > 0 else 0.0,
                        },
                        'objects': []
                    }

                    for obj_file, obj_data in self.parsed_data[library_file].items():
                        flash_size, ram_size, text_size, dec_size, hex_size = self._calculate_sizes(
                            obj_data['text'], obj_data['rodata'], obj_data['data'], obj_data['bss']
                        )
                        library_info['objects'].append({
                            'name': obj_file,
                            'sizes': {
                                'flash': flash_size,
                                'ram': ram_size,
                                'text': text_size,
                                'code': obj_data['text'],
                                'rodata': obj_data['rodata'],
                                'data': obj_data['data'],
                                'bss': obj_data['bss'],
                                'dec': dec_size,
                                'hex': hex_size,
                            },
                            'percentage': {
                                'flash': (flash_size / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0,
                                'ram': (ram_size / self.summary['total_ram'] * 100) if self.summary['total_ram'] > 0 else 0.0,
                            }
                        })

                    json_data['libraries'].append(library_info)

                with open(json_path, 'w', encoding='utf-8') as f_json:
                    json.dump(json_data, f_json, indent=2, ensure_ascii=False)

            except IOError as e:
                raise IOError(f'Error writing JSON file: {e}') from e

    def print_summary(self) -> None:
        """Print summary statistics to console."""
        self._calculate_summary()
        print('\n' + '=' * 80)
        print('Size Map Summary')
        print('=' * 80)
        print(f'Total Flash: {self.summary["total_flash"]:,} bytes ({self.summary["total_flash"]/1024:.2f} KB)')
        print(f'Total RAM:   {self.summary["total_ram"]:,} bytes ({self.summary["total_ram"]/1024:.2f} KB)')
        print(f'Total Code:  {self.summary["total_code"]:,} bytes ({self.summary["total_code"]/1024:.2f} KB)')
        print(f'Total Rodata: {self.summary["total_rodata"]:,} bytes ({self.summary["total_rodata"]/1024:.2f} KB)')
        print(f'Total Data:  {self.summary["total_data"]:,} bytes ({self.summary["total_data"]/1024:.2f} KB)')
        print(f'Total BSS:   {self.summary["total_bss"]:,} bytes ({self.summary["total_bss"]/1024:.2f} KB)')
        print(f'Libraries:  {self.summary["library_count"]}')
        print(f'Objects:    {self.summary["object_count"]}')
        print('=' * 80 + '\n')

        # Print top 10 largest libraries
        sorted_libraries = self._get_sorted_libraries()
        if sorted_libraries:
            print('Top 10 Largest Libraries (by Flash):')
            print('-' * 80)
            for i, (library_file, totals) in enumerate(sorted_libraries[:10], 1):
                percentage = (totals['flash'] / self.summary['total_flash'] * 100) if self.summary['total_flash'] > 0 else 0.0
                print(f'{i:2d}. {os.path.basename(library_file):50s} '
                      f'{totals["flash"]:8,} bytes ({percentage:5.2f}%)')
            print()

    @staticmethod
    def compare_maps(old_map: 'MapHelper', new_map: 'MapHelper',
                     output_dir: str = None, output_basename: str = None) -> Dict:
        """
        Compare two map files and generate difference report.

        Args:
            old_map: MapHelper instance for the old/baseline map file
            new_map: MapHelper instance for the new/current map file
            output_dir: Output directory for comparison reports
            output_basename: Base name for output files

        Returns:
            Dictionary containing comparison results
        """
        old_map._calculate_summary()
        new_map._calculate_summary()

        if output_dir is None:
            output_dir = os.path.dirname(new_map.map_file_path)
        if output_basename is None:
            output_basename = 'map_comparison'

        # Compare library-level data
        old_libraries = old_map._get_sorted_libraries()
        new_libraries = new_map._get_sorted_libraries()

        old_lib_dict = {lib_file: totals for lib_file, totals in old_libraries}
        new_lib_dict = {lib_file: totals for lib_file, totals in new_libraries}

        # Find added, removed, and modified libraries
        old_lib_set = set(old_lib_dict.keys())
        new_lib_set = set(new_lib_dict.keys())

        added_libs = new_lib_set - old_lib_set
        removed_libs = old_lib_set - new_lib_set
        common_libs = old_lib_set & new_lib_set

        # Compare libraries
        lib_changes = []
        for lib_file in common_libs:
            old_totals = old_lib_dict[lib_file]
            new_totals = new_lib_dict[lib_file]

            changes = {}
            for key in ['flash', 'ram', 'text', 'code', 'rodata', 'data', 'bss']:
                old_val = old_totals.get(key, 0)
                new_val = new_totals.get(key, 0)
                diff = new_val - old_val
                if diff != 0:
                    changes[key] = {
                        'old': old_val,
                        'new': new_val,
                        'diff': diff,
                        'diff_pct': (diff / old_val * 100) if old_val > 0 else (100 if new_val > 0 else 0)
                    }

            if changes:
                lib_changes.append({
                    'library': lib_file,
                    'changes': changes
                })

        # Compare object-level data
        obj_changes = []
        obj_added = []
        obj_removed = []

        # Compare objects in common libraries
        for lib_file in common_libs:
            old_objs = old_map.parsed_data.get(lib_file, {})
            new_objs = new_map.parsed_data.get(lib_file, {})

            old_obj_set = set(old_objs.keys())
            new_obj_set = set(new_objs.keys())

            # Added objects
            for obj_name in new_obj_set - old_obj_set:
                obj_data = new_objs[obj_name]
                flash_size, ram_size, text_size, dec_size, hex_size = new_map._calculate_sizes(
                    obj_data.get('text', 0), obj_data.get('rodata', 0),
                    obj_data.get('data', 0), obj_data.get('bss', 0)
                )
                obj_added.append({
                    'library': lib_file,
                    'object': obj_name,
                    'flash': flash_size,
                    'ram': ram_size,
                    'text': text_size,
                    'code': obj_data.get('text', 0),
                    'rodata': obj_data.get('rodata', 0),
                    'data': obj_data.get('data', 0),
                    'bss': obj_data.get('bss', 0),
                })

            # Removed objects
            for obj_name in old_obj_set - new_obj_set:
                obj_data = old_objs[obj_name]
                flash_size, ram_size, text_size, dec_size, hex_size = old_map._calculate_sizes(
                    obj_data.get('text', 0), obj_data.get('rodata', 0),
                    obj_data.get('data', 0), obj_data.get('bss', 0)
                )
                obj_removed.append({
                    'library': lib_file,
                    'object': obj_name,
                    'flash': flash_size,
                    'ram': ram_size,
                    'text': text_size,
                    'code': obj_data.get('text', 0),
                    'rodata': obj_data.get('rodata', 0),
                    'data': obj_data.get('data', 0),
                    'bss': obj_data.get('bss', 0),
                })

            # Modified objects
            for obj_name in old_obj_set & new_obj_set:
                old_obj_data = old_objs[obj_name]
                new_obj_data = new_objs[obj_name]

                old_flash, old_ram, old_text, _, _ = old_map._calculate_sizes(
                    old_obj_data.get('text', 0), old_obj_data.get('rodata', 0),
                    old_obj_data.get('data', 0), old_obj_data.get('bss', 0)
                )
                new_flash, new_ram, new_text, _, _ = new_map._calculate_sizes(
                    new_obj_data.get('text', 0), new_obj_data.get('rodata', 0),
                    new_obj_data.get('data', 0), new_obj_data.get('bss', 0)
                )

                changes = {}
                for key in ['text', 'rodata', 'data', 'bss']:
                    old_val = old_obj_data.get(key, 0)
                    new_val = new_obj_data.get(key, 0)
                    diff = new_val - old_val
                    if diff != 0:
                        changes[key] = {
                            'old': old_val,
                            'new': new_val,
                            'diff': diff
                        }

                if changes or old_flash != new_flash or old_ram != new_ram:
                    obj_changes.append({
                        'library': lib_file,
                        'object': obj_name,
                        'flash': {'old': old_flash, 'new': new_flash, 'diff': new_flash - old_flash},
                        'ram': {'old': old_ram, 'new': new_ram, 'diff': new_ram - old_ram},
                        'text': {'old': old_text, 'new': new_text, 'diff': new_text - old_text},
                        'changes': changes
                    })

        # Calculate summary differences
        summary_diff = {}
        for key in ['total_flash', 'total_ram', 'total_text', 'total_code',
                   'total_rodata', 'total_data', 'total_bss']:
            old_val = old_map.summary.get(key, 0)
            new_val = new_map.summary.get(key, 0)
            diff = new_val - old_val
            summary_diff[key] = {
                'old': old_val,
                'new': new_val,
                'diff': diff,
                'diff_pct': (diff / old_val * 100) if old_val > 0 else (100 if new_val > 0 else 0)
            }

        comparison_result = {
            'old_file': old_map.map_file_path,
            'new_file': new_map.map_file_path,
            'summary_diff': summary_diff,
            'added_libraries': [{'library': lib, 'totals': new_lib_dict[lib]} for lib in added_libs],
            'removed_libraries': [{'library': lib, 'totals': old_lib_dict[lib]} for lib in removed_libs],
            'modified_libraries': lib_changes,
            'added_objects': obj_added,
            'removed_objects': obj_removed,
            'modified_objects': obj_changes,
        }

        # Generate comparison reports
        MapHelper._save_comparison_reports(comparison_result, output_dir, output_basename)

        return comparison_result

    @staticmethod
    def _save_comparison_reports(comparison: Dict, output_dir: str, output_basename: str) -> None:
        """Save comparison reports in multiple formats."""
        # Save text report
        txt_path = os.path.join(output_dir, f'{output_basename}.txt')
        MapHelper._save_comparison_text(comparison, txt_path)

        # Save CSV reports
        csv_summary_path = os.path.join(output_dir, f'{output_basename}_summary.csv')
        csv_lib_path = os.path.join(output_dir, f'{output_basename}_libraries.csv')
        csv_obj_path = os.path.join(output_dir, f'{output_basename}_objects.csv')
        MapHelper._save_comparison_csv(comparison, csv_summary_path, csv_lib_path, csv_obj_path)

        # Save Excel report
        if EXCEL_SUPPORT:
            excel_path = os.path.join(output_dir, f'{output_basename}.xlsx')
            MapHelper._save_comparison_excel(comparison, excel_path)

    @staticmethod
    def _save_comparison_text(comparison: Dict, txt_path: str) -> None:
        """
        Save text format comparison report with detailed library and object changes.

        The text report includes:
        - Summary differences (Flash, RAM, Code, Rodata, Data, BSS)
        - Added/Removed/Modified libraries with size information
        - Added/Removed/Modified objects (.c.obj files) with detailed size changes
        - For modified objects, shows Flash/RAM/Text old/new/diff values and segment-level changes
        """
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.write('=' * 80 + '\n')
            f.write('Map File Comparison Report\n')
            f.write('=' * 80 + '\n')
            f.write(f'Old File: {comparison["old_file"]}\n')
            f.write(f'New File: {comparison["new_file"]}\n')
            f.write('=' * 80 + '\n\n')

            # Summary differences
            f.write('SUMMARY DIFFERENCES\n')
            f.write('-' * 80 + '\n')
            f.write(f'{"Metric":<20} {"Old":>15} {"New":>15} {"Diff":>15} {"Diff %":>10}\n')
            f.write('-' * 80 + '\n')
            for key, diff in comparison['summary_diff'].items():
                metric = key.replace('total_', '').upper()
                diff_str = f'{diff["diff"]:>+15,}'
                diff_pct_str = f'{diff["diff_pct"]:>+9.2f}%'
                f.write(f'{metric:<20} {diff["old"]:>15,} {diff["new"]:>15,} '
                       f'**{diff_str.strip()}** **{diff_pct_str.strip()}**\n')
            f.write('\n')

            # Added libraries
            if comparison['added_libraries']:
                f.write('ADDED LIBRARIES\n')
                f.write('-' * 80 + '\n')
                for item in comparison['added_libraries']:
                    lib_name = os.path.basename(item['library'])
                    totals = item['totals']
                    f.write(f'{lib_name}: Flash={totals["flash"]:,}, RAM={totals["ram"]:,}\n')
                f.write('\n')

            # Removed libraries
            if comparison['removed_libraries']:
                f.write('REMOVED LIBRARIES\n')
                f.write('-' * 80 + '\n')
                for item in comparison['removed_libraries']:
                    lib_name = os.path.basename(item['library'])
                    totals = item['totals']
                    f.write(f'{lib_name}: Flash={totals["flash"]:,}, RAM={totals["ram"]:,}\n')
                f.write('\n')

            # Modified libraries
            if comparison['modified_libraries']:
                f.write('MODIFIED LIBRARIES\n')
                f.write('-' * 80 + '\n')
                for item in comparison['modified_libraries']:
                    lib_name = os.path.basename(item['library'])
                    f.write(f'{lib_name}:\n')
                    for key, change in item['changes'].items():
                        diff_value = f'{change["diff"]:+,}'
                        diff_pct_value = f'{change["diff_pct"]:+.2f}%'
                        f.write(f'  {key}: {change["old"]:,} -> {change["new"]:,} '
                               f'(**{diff_value}**, **{diff_pct_value}**)\n')
                f.write('\n')

            # Object changes summary
            f.write('OBJECT CHANGES SUMMARY\n')
            f.write('-' * 80 + '\n')
            f.write(f'Added Objects:   {len(comparison["added_objects"])}\n')
            f.write(f'Removed Objects: {len(comparison["removed_objects"])}\n')
            f.write(f'Modified Objects: {len(comparison["modified_objects"])}\n')
            f.write('\n')

            # Added objects detail
            if comparison['added_objects']:
                f.write('ADDED OBJECTS\n')
                f.write('-' * 80 + '\n')
                for item in comparison['added_objects']:
                    lib_name = os.path.basename(item['library'])
                    obj_name = item['object']
                    f.write(f'{lib_name}({obj_name}): Flash={item["flash"]:,}, RAM={item["ram"]:,}, '
                           f'Text={item["text"]:,}\n')
                f.write('\n')

            # Removed objects detail
            if comparison['removed_objects']:
                f.write('REMOVED OBJECTS\n')
                f.write('-' * 80 + '\n')
                for item in comparison['removed_objects']:
                    lib_name = os.path.basename(item['library'])
                    obj_name = item['object']
                    f.write(f'{lib_name}({obj_name}): Flash={item["flash"]:,}, RAM={item["ram"]:,}, '
                           f'Text={item["text"]:,}\n')
                f.write('\n')

            # Modified objects detail
            if comparison['modified_objects']:
                f.write('MODIFIED OBJECTS\n')
                f.write('-' * 80 + '\n')
                for item in comparison['modified_objects']:
                    lib_name = os.path.basename(item['library'])
                    obj_name = item['object']
                    f.write(f'{lib_name}({obj_name}):\n')
                    flash_diff = f'{item["flash"]["diff"]:+,}'
                    ram_diff = f'{item["ram"]["diff"]:+,}'
                    text_diff = f'{item["text"]["diff"]:+,}'
                    f.write(f'  Flash: {item["flash"]["old"]:,} -> {item["flash"]["new"]:,} '
                           f'(**{flash_diff}**)\n')
                    f.write(f'  RAM:   {item["ram"]["old"]:,} -> {item["ram"]["new"]:,} '
                           f'(**{ram_diff}**)\n')
                    f.write(f'  Text:  {item["text"]["old"]:,} -> {item["text"]["new"]:,} '
                           f'(**{text_diff}**)\n')
                    if item['changes']:
                        f.write('  Segment changes:\n')
                        for key, change in item['changes'].items():
                            change_diff = f'{change["diff"]:+,}'
                            f.write(f'    {key}: {change["old"]:,} -> {change["new"]:,} '
                                   f'(**{change_diff}**)\n')
                    f.write('\n')

    @staticmethod
    def _save_comparison_csv(comparison: Dict, summary_path: str, lib_path: str, obj_path: str) -> None:
        """Save CSV format comparison reports."""
        # Summary CSV
        with open(summary_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['metric', 'old', 'new', 'diff', 'diff_pct'])
            for key, diff in comparison['summary_diff'].items():
                metric = key.replace('total_', '')
                writer.writerow([metric, diff['old'], diff['new'], diff['diff'], f'{diff["diff_pct"]:.2f}'])

        # Libraries CSV
        with open(lib_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['status', 'library', 'flash_old', 'flash_new', 'flash_diff',
                           'ram_old', 'ram_new', 'ram_diff', 'text_old', 'text_new', 'text_diff'])

            for item in comparison['added_libraries']:
                totals = item['totals']
                writer.writerow(['added', item['library'], 0, totals['flash'], totals['flash'],
                               0, totals['ram'], totals['ram'], 0, totals['text'], totals['text']])

            for item in comparison['removed_libraries']:
                totals = item['totals']
                writer.writerow(['removed', item['library'], totals['flash'], 0, -totals['flash'],
                               totals['ram'], 0, -totals['ram'], totals['text'], 0, -totals['text']])

            for item in comparison['modified_libraries']:
                lib = item['library']
                changes = item['changes']
                flash_old = changes.get('flash', {}).get('old', 0)
                flash_new = changes.get('flash', {}).get('new', 0)
                ram_old = changes.get('ram', {}).get('old', 0)
                ram_new = changes.get('ram', {}).get('new', 0)
                text_old = changes.get('text', {}).get('old', 0)
                text_new = changes.get('text', {}).get('new', 0)
                writer.writerow(['modified', lib, flash_old, flash_new, flash_new - flash_old,
                               ram_old, ram_new, ram_new - ram_old, text_old, text_new, text_new - text_old])

        # Objects CSV
        with open(obj_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['status', 'library', 'object', 'flash_old', 'flash_new', 'flash_diff',
                           'ram_old', 'ram_new', 'ram_diff', 'text_old', 'text_new', 'text_diff'])

            for item in comparison['added_objects']:
                writer.writerow(['added', item['library'], item['object'], 0, item['flash'], item['flash'],
                               0, item['ram'], item['ram'], 0, item['text'], item['text']])

            for item in comparison['removed_objects']:
                writer.writerow(['removed', item['library'], item['object'], item['flash'], 0, -item['flash'],
                               item['ram'], 0, -item['ram'], item['text'], 0, -item['text']])

            for item in comparison['modified_objects']:
                writer.writerow(['modified', item['library'], item['object'],
                               item['flash']['old'], item['flash']['new'], item['flash']['diff'],
                               item['ram']['old'], item['ram']['new'], item['ram']['diff'],
                               item['text']['old'], item['text']['new'], item['text']['diff']])

    @staticmethod
    def _save_comparison_excel(comparison: Dict, excel_path: str) -> None:
        """Save Excel format comparison report with multiple sheets."""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Summary sheet
        ws_summary = wb.create_sheet('Summary', 0)
        ws_summary.append(['Metric', 'Old', 'New', 'Diff', 'Diff %'])
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for key, diff in comparison['summary_diff'].items():
            metric = key.replace('total_', '').upper()
            ws_summary.append([metric, diff['old'], diff['new'], diff['diff'], f'{diff["diff_pct"]:.2f}%'])

        # Libraries sheet
        ws_libs = wb.create_sheet('Libraries', 1)
        ws_libs.append(['Status', 'Library', 'Flash Old', 'Flash New', 'Flash Diff',
                       'RAM Old', 'RAM New', 'RAM Diff', 'Text Old', 'Text New', 'Text Diff'])
        for cell in ws_libs[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for item in comparison['added_libraries']:
            totals = item['totals']
            ws_libs.append(['Added', item['library'], 0, totals['flash'], totals['flash'],
                          0, totals['ram'], totals['ram'], 0, totals['text'], totals['text']])

        for item in comparison['removed_libraries']:
            totals = item['totals']
            ws_libs.append(['Removed', item['library'], totals['flash'], 0, -totals['flash'],
                          totals['ram'], 0, -totals['ram'], totals['text'], 0, -totals['text']])

        for item in comparison['modified_libraries']:
            lib = item['library']
            changes = item['changes']
            flash_old = changes.get('flash', {}).get('old', 0)
            flash_new = changes.get('flash', {}).get('new', 0)
            ram_old = changes.get('ram', {}).get('old', 0)
            ram_new = changes.get('ram', {}).get('new', 0)
            text_old = changes.get('text', {}).get('old', 0)
            text_new = changes.get('text', {}).get('new', 0)
            ws_libs.append(['Modified', lib, flash_old, flash_new, flash_new - flash_old,
                          ram_old, ram_new, ram_new - ram_old, text_old, text_new, text_new - text_old])

        # Objects sheet
        ws_objs = wb.create_sheet('Objects', 2)
        ws_objs.append(['Status', 'Library', 'Object', 'Flash Old', 'Flash New', 'Flash Diff',
                       'RAM Old', 'RAM New', 'RAM Diff', 'Text Old', 'Text New', 'Text Diff'])
        for cell in ws_objs[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for item in comparison['added_objects']:
            ws_objs.append(['Added', item['library'], item['object'], 0, item['flash'], item['flash'],
                          0, item['ram'], item['ram'], 0, item['text'], item['text']])

        for item in comparison['removed_objects']:
            ws_objs.append(['Removed', item['library'], item['object'], item['flash'], 0, -item['flash'],
                          item['ram'], 0, -item['ram'], item['text'], 0, -item['text']])

        for item in comparison['modified_objects']:
            ws_objs.append(['Modified', item['library'], item['object'],
                          item['flash']['old'], item['flash']['new'], item['flash']['diff'],
                          item['ram']['old'], item['ram']['new'], item['ram']['diff'],
                          item['text']['old'], item['text']['new'], item['text']['diff']])

        # Auto-adjust column widths and freeze panes
        for ws in [ws_summary, ws_libs, ws_objs]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            ws.freeze_panes = 'A2'

        wb.save(excel_path)


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='''Parse linker map file and generate detailed size statistics reports.

This tool analyzes linker-generated .map files and produces comprehensive size
statistics for Flash and RAM usage, broken down by library and object files.

OVERVIEW:
  The tool parses ARM GCC linker map files (.map) and extracts size information
  for each library and object file. It categorizes sizes into Flash (code + rodata + data)
  and RAM (data + bss) usage, providing detailed breakdowns for optimization analysis.

OUTPUT FILES:
  The tool generates output files based on the input map file name to avoid overwriting
  when processing multiple map files:

  - <input_name>_size_map.txt
      Human-readable text format with summary statistics at the top, followed by
      detailed breakdown per library and object file.

  - <input_name>_size_map.xlsx
      Excel file with multiple sheets:
      * "Total" sheet: Summary statistics per library file (.a files)
      * "Detail" sheet: Detailed statistics per object file (.obj files)
      * "Top15_Text" sheet: Top 15 libraries by text size (code+rodata) with stacked chart
      * "Top15_BSS" sheet: Top 15 libraries by BSS size with bar chart
      * "Top15_Data" sheet: Top 15 libraries by Data (rwdata) size with bar chart
      * "Top15_RAM" sheet: Top 15 libraries by RAM size (data+bss) with stacked chart
      * "SortedByFlashSize"/"SortedByFlashSizeDetail": Flash-descending replicas of the CSV sheets including a `rank` column
      * "SortedByLibraryName"/"SortedByLibraryNameDetail": Same data sorted by library name ascending
      This is the recommended format for viewing and analysis.

  - <input_name>_size_map_total.csv
      Summary statistics per library file (.a files). One row per library with
      aggregated sizes for all objects in that library. `filename` now contains
      only `lib***.a`, and `library_path` is the final column, so the archive
      directory is preserved once but not repeated. (Also available in the Excel
      "Total" and sorted sheets.)

  - <input_name>_size_map_detail.csv
      Detailed statistics per object file (.obj files). One row per object file
      with individual sizes. `obj_name` is paired with the basename `filename`,
      and the final `library_path` column keeps the archive path for reference.
      (Also available in the Excel "Detail" and sorted sheets.)

  - <input_name>_size_map.json
      Complete data in JSON format with percentages. Suitable for programmatic
      analysis, integration with CI/CD tools, or generating custom reports.

SIZE CATEGORIES:
  The tool calculates and reports the following size categories:

  - flash:  Total Flash usage (code + rodata + data)
            This is what gets stored in Flash memory.

  - ram:    Total RAM usage (data + bss)
            This is what gets loaded into RAM at runtime.

  - text:   Text segment size (code + rodata)
            Read-only executable code and constants.

  - code:   Code segment size only
            Executable instructions (.text sections).

  - rodata: Read-only data segment size
            Constants, strings, and other read-only data (.rodata sections).

  - data:   Initialized data segment size
            Global/static variables with initial values (.data sections).

  - bss:    Uninitialized data segment size
            Global/static variables initialized to zero (.bss sections).

  - dec:    Total size in decimal (text + data + bss)

  - hex:    Total size in hexadecimal format

COMMAND LINE EXAMPLES:

  1. Basic usage - Generate all output formats (TXT, CSV, JSON, Excel):
     %(prog)s build/bk7236n/app/bk7236n/app.map

     Output files:
       - app_size_map.txt
       - app_size_map.xlsx (Total, Detail, Top15_Text, Top15_BSS, Top15_Data, Top15_RAM sheets with charts)
       - app_size_map_total.csv
       - app_size_map_detail.csv
       - app_size_map.json

  2. Filter large libraries (only show libraries > 10KB Flash):
     %(prog)s app.map --min-size 10240

     This filters out small libraries and focuses on major size consumers.

  3. Sort by RAM usage instead of Flash:
     %(prog)s app.map --sort-by ram

     Useful for identifying RAM-intensive modules.

  4. Generate only JSON format for programmatic analysis:
     %(prog)s app.map --format json

     Output: app_size_map.json only

  5. View summary statistics without generating files:
     %(prog)s app.map --summary --no-output

     Prints summary to console including:
       - Total Flash/RAM usage
       - Top 10 largest libraries
       - Library and object counts

  6. Generate CSV files sorted by code size:
     %(prog)s app.map --format csv --sort-by code

     Output: CSV files only, sorted by code segment size

  7. Filter and analyze large Flash consumers:
     %(prog)s app.map --min-size 50000 --sort-by flash --summary

     Shows only libraries > 50KB, sorted by Flash usage, with summary

  8. Process multiple map files at once (each generates separate outputs):
     %(prog)s build/app/app.map build/bootloader/bootloader.map

     Results:
       - app_size_map_*.csv (from app.map)
       - bootloader_size_map_*.csv (from bootloader.map)
       No overwriting! Each file generates its own output files.

  8a. Process multiple map files (automatically generates merged Excel with separate sheets):
     %(prog)s app.map bootloader.map

     Results:
       - Individual output files for each map file (as above)
       - merged_size_map.xlsx (automatically generated, contains separate sheets:
         app_Total, app_Detail, app_Top15_Text, app_Top15_BSS, app_Top15_Data, app_Top15_RAM,
         bootloader_Total, bootloader_Detail, bootloader_Top15_Text, bootloader_Top15_BSS,
         bootloader_Top15_Data, bootloader_Top15_RAM)
       Each map file's data is kept in separate sheets, not merged together.
       Each Top15 sheet includes visualization charts.

  9. Quick analysis - Summary only for quick overview:
     %(prog)s app.map --no-output

     Fast way to see total sizes and top consumers without file I/O

  10. Integration with build system:
      In CMake or Makefile:
        python3 %(prog)s $(BUILD_DIR)/app.map --format json

      Then parse JSON in CI/CD pipeline for size tracking

  11. Compare two map files to see size differences:
     %(prog)s new_build/app.map --compare old_build/app.map

     Generates comparison reports showing:
     - Summary differences (Flash, RAM, Code, Rodata, Data, BSS)
     - Added/Removed/Modified libraries
     - Added/Removed/Modified objects
     - Detailed changes with old/new values and differences

     Output files:
       - <old>_vs_<new>_comparison.txt (text report with detailed library and object changes)
       - <old>_vs_<new>_comparison.xlsx (Excel with Summary, Libraries, Objects sheets)
       - <old>_vs_<new>_comparison_summary.csv
       - <old>_vs_<new>_comparison_libraries.csv
       - <old>_vs_<new>_comparison_objects.csv

     The text report includes detailed lists of:
       - All added/removed/modified libraries with size information
       - All added/removed/modified objects (.c.obj files) with Flash/RAM/Text sizes
       - For modified objects, shows old/new/diff values and segment-level changes (text/rodata/data/bss)

TIPS AND BEST PRACTICES:

  • Use --min-size to filter out noise and focus on major consumers
    Example: --min-size 10240 filters libraries smaller than 10KB

  • Use --sort-by to identify largest consumers of specific resources
    - flash:  Find what's consuming Flash space
    - ram:    Find what's consuming RAM
    - code:   Find code-heavy modules
    - rodata: Find constant/data-heavy modules

  • Use --format json for:
    - Integration with CI/CD tools
    - Automated size tracking over time
    - Custom report generation
    - Comparison between builds

  • Use --compare to track size changes:
    Compare two builds to see what changed
    Identify size regressions or improvements
    Track library and object-level changes
    Example: %(prog)s new.map --compare old.map

  • Process multiple files efficiently:
    Specify multiple map files to process at once
    Each file generates separate output (no overwriting)
    Automatic merged Excel file for easy comparison
    Useful for comparing different build configurations

  • Use --summary for quick overview without opening files
    Perfect for build logs or quick checks

  • Output file names are based on input file name
    This allows processing multiple map files without overwriting

  • Combine options for powerful analysis:
    --min-size 50000 --sort-by ram --summary
    Shows large RAM consumers with summary

DEPENDENCIES:

  • Excel support (optional):
    The tool can generate Excel files (.xlsx) if openpyxl is installed.
    If not installed, Excel generation will be skipped with a warning.
    Install with: pip install openpyxl
    Note: CSV and other formats work without openpyxl.

TROUBLESHOOTING:

  • If map file is not found:
    Check the path and ensure the file exists
    Error: Map file does not exist: <path>

  • If output files are not generated:
    Check write permissions in the output directory
    Ensure sufficient disk space

  • If Excel files are not generated:
    Install openpyxl: pip install openpyxl
    The tool will continue to work with other formats (CSV, TXT, JSON)

  • If parsing fails:
    Verify the map file is from ARM GCC linker
    Check file encoding (should be UTF-8 compatible)

  • If comparison shows unexpected results:
    Ensure both map files are from the same linker version
    Check that both files are complete and not corrupted
    Verify that the map files use the same memory layout''',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
EXAMPLES (Command Line Demos):

  Example 1: Basic usage - Generate all output formats:
    $ python3 map_helper.py build/bk7236n/app/bk7236n/app.map
    Output: app_size_map.txt, app_size_map.xlsx (6 sheets: Total, Detail, Top15_Text, Top15_BSS, Top15_Data, Top15_RAM),
            app_size_map_total.csv, app_size_map_detail.csv, app_size_map.json

  Example 1a: Generate only Excel file (recommended for viewing):
    $ python3 map_helper.py app.map --format excel
    Output: app_size_map.xlsx with 6 sheets including visualization charts for Top15 library analysis

  Example 2: Filter large libraries (> 10KB Flash):
    $ python3 map_helper.py app.map --min-size 10240
    Only shows libraries consuming more than 10KB Flash space

  Example 3: Sort by RAM usage:
    $ python3 map_helper.py app.map --sort-by ram
    Identifies RAM-intensive modules

  Example 4: Generate only JSON format:
    $ python3 map_helper.py app.map --format json
    Output: app_size_map.json only

  Example 5: Quick summary without files:
    $ python3 map_helper.py app.map --summary --no-output
    Prints statistics to console, no files generated

  Example 6: CSV files sorted by code size:
    $ python3 map_helper.py app.map --format csv --sort-by code
    Output: CSV files only, sorted by code segment

  Example 7: Analyze large Flash consumers:
    $ python3 map_helper.py app.map --min-size 50000 --sort-by flash --summary
    Shows libraries > 50KB, sorted by Flash, with summary

  Example 8: Process multiple map files at once:
    $ python3 map_helper.py build/app/app.map build/bootloader/bootloader.map
    Each generates separate output files (no overwriting)

  Example 8a: Process multiple map files (auto-generates merged Excel with separate sheets):
    $ python3 map_helper.py app.map bootloader.map
    Generates individual files + merged_size_map.xlsx with separate sheets for each file
    Each map file has its own Total, Detail, Top15_Text, Top15_BSS, Top15_Data, Top15_RAM sheets (data not merged)
    All Top15 sheets include visualization charts for quick hotspot analysis

  Example 9: Quick analysis:
    $ python3 map_helper.py app.map --no-output
    Fast overview without file I/O

  Example 10: Integration with build system:
    $ python3 map_helper.py $(BUILD_DIR)/app.map --format json
    Use in CMake/Makefile for CI/CD integration

  Example 11: Compare two map files to see what changed:
    $ python3 map_helper.py new_build/app.map --compare old_build/app.map
    Generates detailed comparison showing size differences, added/removed/modified
    libraries and objects, with old/new values and percentage changes.
    The text report lists all changed .c.obj files with detailed size information.

For more information, see the source code or documentation.
        ''',
        add_help=True
    )
    parser.add_argument('map_file', metavar='MAP_FILE', nargs='+',
                        help='Path(s) to the linker-generated .map file(s) to analyze. '
                             'You can specify multiple map files to process them all at once. '
                             'Each file will generate separate output files based on its name.')
    parser.add_argument('--min-size', type=int, default=0, metavar='BYTES',
                        help='''Minimum size threshold for filtering libraries (in bytes).
Libraries smaller than this threshold will be excluded from output.
Useful for focusing on major size consumers. Default: 0 (no filtering).
Examples: 1024 (1KB), 10240 (10KB), 1048576 (1MB)''')
    parser.add_argument('--sort-by', choices=['flash', 'ram', 'text', 'code', 'rodata', 'data', 'bss', 'dec'],
                        default='flash', metavar='FIELD',
                        help='''Sort libraries by this field in descending order.
Available fields:
  flash  - Total Flash usage (code + rodata + data) [default]
  ram    - Total RAM usage (data + bss)
  text   - Text segment size (code + rodata)
  code   - Code segment size only
  rodata - Read-only data segment size
  data   - Initialized data segment size
  bss    - Uninitialized data segment size
  dec    - Total size in decimal''')
    parser.add_argument('--format', choices=['all', 'txt', 'csv', 'json', 'excel'],
                        default='all', metavar='FORMAT',
                        help='''Output format to generate.
  all   - Generate all formats (TXT, CSV, JSON, Excel) [default]
  txt   - Generate only text format (.txt)
  csv   - Generate CSV files (.csv) and Excel file (.xlsx)
  excel - Generate only Excel file (.xlsx) with two sheets
  json  - Generate only JSON format (.json)
Note: Excel file contains 6 sheets: "Total" (library summary), "Detail" (object details),
      "Top15_Text" (top 15 libraries by text), "Top15_BSS" (top 15 libraries by BSS),
      "Top15_Data" (top 15 libraries by data), "Top15_RAM" (top 15 libraries by RAM).
      All Top15 sheets include visualization charts.
      CSV files are separate files. Output file names are based on input map file name.''')
    parser.add_argument('--no-excel', action='store_true',
                        help='''Do not generate Excel file (.xlsx).
By default, Excel file is generated when using --format all or csv.
Use this flag to skip Excel generation and only create CSV files.''')
    parser.add_argument('--summary', action='store_true',
                        help='''Print detailed summary statistics to console.
Shows total Flash/RAM usage, library/object counts, and top 10 largest
libraries with percentages. Useful for quick overview without opening files.''')
    parser.add_argument('--no-output', action='store_true',
                        help='''Do not generate any output files.
Only print summary statistics to console. Useful for quick analysis
or when you only need console output. Implies --summary.''')
    parser.add_argument('--merge', action='store_true',
                        help='''[Deprecated] When processing multiple map files, a merged Excel file is automatically generated.
Each map file will have its own "Total" and "Detail" sheets named after the input file.
Data from different map files are kept in separate sheets (not merged into one sheet).
Example: Processing app.map and bootloader.map will create sheets: "app_Total", "app_Detail", "bootloader_Total", "bootloader_Detail"''')
    parser.add_argument('--compare', metavar='OLD_MAP',
                        help='''Compare two map files and generate difference report.
Specify the old/baseline map file path. The first map file in the argument list will be used as the new/current file.
This mode generates detailed comparison reports showing:
  - Summary differences (Flash, RAM, Code, Rodata, Data, BSS)
  - Added/Removed/Modified libraries with size changes
  - Added/Removed/Modified objects with size changes
  - Old/new values and percentage changes for all metrics

Output files:
  - <old>_vs_<new>_comparison.txt (text report with detailed library and object changes)
  - <old>_vs_<new>_comparison.xlsx (Excel with Summary, Libraries, Objects sheets)
  - <old>_vs_<new>_comparison_summary.csv
  - <old>_vs_<new>_comparison_libraries.csv
  - <old>_vs_<new>_comparison_objects.csv

The text report includes detailed lists of all changed libraries and objects (.c.obj files),
showing Flash/RAM/Text sizes and segment-level changes for easy identification of what changed.

Example: %(prog)s new_build/app.map --compare old_build/app.map
This will generate comparison reports showing what changed between old.map and new.map.''')

    args = parser.parse_args()

    # Handle comparison mode
    if args.compare:
        if len(args.map_file) != 1:
            print('Error: When using --compare, specify exactly one new map file.', file=sys.stderr)
            sys.exit(1)

        old_map_path = args.compare
        new_map_path = args.map_file[0]

        try:
            print(f'\n{"=" * 80}')
            print('Map File Comparison Mode')
            print('=' * 80)
            print(f'Old/Baseline file: {old_map_path}')
            print(f'New/Current file:  {new_map_path}')
            print('=' * 80)

            # Parse both map files
            old_map = MapHelper(old_map_path, min_size=args.min_size, sort_by=args.sort_by)
            old_map.start_parse()

            new_map = MapHelper(new_map_path, min_size=args.min_size, sort_by=args.sort_by)
            new_map.start_parse()

            # Generate comparison
            output_dir = os.path.dirname(new_map_path)
            old_basename = os.path.basename(old_map_path)
            new_basename = os.path.basename(new_map_path)
            if old_basename.endswith('.map'):
                old_basename = old_basename[:-4]
            if new_basename.endswith('.map'):
                new_basename = new_basename[:-4]
            output_basename = f'{old_basename}_vs_{new_basename}_comparison'

            comparison = MapHelper.compare_maps(old_map, new_map, output_dir, output_basename)

            # Print summary to console
            print('\n' + '=' * 80)
            print('COMPARISON SUMMARY')
            print('=' * 80)
            print(f'{"Metric":<20} {"Old":>15} {"New":>15} {"Diff":>15} {"Diff %":>10}')
            print('-' * 80)
            for key, diff in comparison['summary_diff'].items():
                metric = key.replace('total_', '').upper()
                print(f'{metric:<20} {diff["old"]:>15,} {diff["new"]:>15,} '
                     f'{diff["diff"]:>+15,} {diff["diff_pct"]:>+9.2f}%')
            print('-' * 80)
            print(f'\nAdded Libraries:   {len(comparison["added_libraries"])}')
            print(f'Removed Libraries: {len(comparison["removed_libraries"])}')
            print(f'Modified Libraries: {len(comparison["modified_libraries"])}')
            print(f'Added Objects:     {len(comparison["added_objects"])}')
            print(f'Removed Objects:   {len(comparison["removed_objects"])}')
            print(f'Modified Objects:  {len(comparison["modified_objects"])}')
            print('=' * 80)

            print(f'\nComparison reports generated in: {output_dir}')
            print(f'  - {output_basename}.txt')
            if EXCEL_SUPPORT:
                print(f'  - {output_basename}.xlsx (Summary, Libraries, Objects sheets)')
            print(f'  - {output_basename}_summary.csv')
            print(f'  - {output_basename}_libraries.csv')
            print(f'  - {output_basename}_objects.csv')

        except FileNotFoundError as e:
            print(f'Error: {e}', file=sys.stderr)
            sys.exit(1)
        except Exception as e:
            print(f'Error: {e}', file=sys.stderr)
            import traceback
            traceback.print_exc()
            sys.exit(1)

        return

    # Normalize map_file to always be a list
    map_files = args.map_file if isinstance(args.map_file, list) else [args.map_file]

    # Process each map file
    processed_files = []
    all_map_helpers = []  # Store all helpers for merge option

    for map_file in map_files:
        try:
            print(f'\n{"=" * 80}')
            print(f'Processing: {map_file}')
            print('=' * 80)

            mh = MapHelper(map_file, min_size=args.min_size, sort_by=args.sort_by)
            mh.start_parse()
            all_map_helpers.append((map_file, mh))

            if args.summary or args.no_output:
                mh.print_summary()

            if not args.no_output:
                excel_format = not args.no_excel and args.format in ('all', 'csv', 'excel')
                mh.save_to_file(output_format=args.format, excel_format=excel_format)
                output_dir = os.path.dirname(map_file)
                output_basename = mh._get_output_basename()
                print(f'Successfully generated size map files in: {output_dir}')
                if args.format in ('all', 'txt'):
                    print(f'  - {output_basename}.txt')
                if args.format in ('all', 'csv'):
                    if not args.no_excel:
                        print(f'  - {output_basename}.xlsx (6 sheets: Total, Detail, Top15_Text, Top15_BSS, Top15_Data, Top15_RAM)')
                    print(f'  - {output_basename}_total.csv')
                    print(f'  - {output_basename}_detail.csv')
                if args.format == 'excel':
                    print(f'  - {output_basename}.xlsx (6 sheets: Total, Detail, Top15_Text, Top15_BSS, Top15_Data, Top15_RAM)')
                if args.format in ('all', 'json'):
                    print(f'  - {output_basename}.json')

                processed_files.append((map_file, output_dir, output_basename))

        except FileNotFoundError as e:
            print(f'Error processing {map_file}: {e}', file=sys.stderr)
            continue
        except Exception as e:
            print(f'Error processing {map_file}: {e}', file=sys.stderr)
            import traceback
            traceback.print_exc()
            continue

    # Generate merged Excel file when processing multiple files (each file has separate sheets)
    # This creates one Excel file with separate sheets for each map file's Total and Detail data
    if len(all_map_helpers) > 1 and not args.no_output and not args.no_excel and args.format in ('all', 'csv', 'excel'):
        if not EXCEL_SUPPORT:
            print('\nWarning: openpyxl not available, cannot generate merged Excel file.', file=sys.stderr)
        else:
            try:
                MapHelper._generate_merged_excel(all_map_helpers, processed_files, args)
            except Exception as e:
                print(f'\nError generating merged Excel file: {e}', file=sys.stderr)
                import traceback
                traceback.print_exc()
    # Also generate merged file if --merge is explicitly requested (for single file case)
    elif args.merge and len(all_map_helpers) == 1 and not args.no_output and not args.no_excel and args.format in ('all', 'csv', 'excel'):
        if not EXCEL_SUPPORT:
            print('\nWarning: openpyxl not available, cannot generate merged Excel file.', file=sys.stderr)
        else:
            try:
                MapHelper._generate_merged_excel(all_map_helpers, processed_files, args)
            except Exception as e:
                print(f'\nError generating merged Excel file: {e}', file=sys.stderr)
                import traceback
                traceback.print_exc()

    # Print summary
    if len(processed_files) > 0:
        print(f'\n{"=" * 80}')
        print(f'Successfully processed {len(processed_files)} map file(s)')
        print('=' * 80)
    elif args.no_output and len(all_map_helpers) > 0:
        print(f'\n{"=" * 80}')
        print(f'Successfully processed {len(all_map_helpers)} map file(s) (--no-output)')
        print('=' * 80)
    elif len(map_files) > 0:
        print('\nNo files were successfully processed.', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()

